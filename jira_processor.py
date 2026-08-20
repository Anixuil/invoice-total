#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Jira 导出数据处理：把 Word 中的手工表格步骤变成可追溯的规则流水线。"""

from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
import re
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidocr_onnxruntime import RapidOCR
import cv2


ProgressCallback = Callable[[dict[str, Any]], None] | None

STATUS_MAP = {
    "已解决": "已完成",
    "已关闭": "已完成",
    "处理中": "进行中",
}
KNOWN_STATUSES = set(STATUS_MAP) | {"开放", "已完成", "进行中"}

# 来源文档中的人员顺序。实际数据中没有任务的人员也保留在汇总中，方便周报直接复制。
DEFAULT_ROSTER = [
    "linyuping",
    "liuxin01",
    "wuyutian",
    "caoyuan",
    "zhengleyuan",
    "douhuanhuan",
    "xiaozhennan",
    "sunyanqiang",
    "linyu01",
    "lilixin",
    "fanshuyu",
    "yinwei",
    "wangruixia",
    "yuhaicheng",
    "caiyuanpeng",
    "ruishunzi",
    "guanhaojie",
    "lihongqi",
    "liwenbo",
    "mojunyou",
    "chenyu",
    "muzhengyi",
]

WEEKLY_ROSTER = [
    ("zhoujiayu", "周佳宇"), ("zhengleyuan", "郑乐园"), ("douhuanhuan", "窦欢欢"),
    ("sunyanqiang", "孙艳强"), ("lilixin", "李礼辛"), ("linyu01", "林榆"),
    ("fanshuyu", "范书毓"), ("xiaozhennan", "肖镇楠"), ("wangruixia", "王瑞霞"),
    ("yinwei", "尹维"), ("yuhaicheng", "余海城"), ("muzhengyi", "穆政逸"),
    ("ruishunzi", "芮顺子"), ("caiyuanpeng", "蔡远鹏"), ("guanhaojie", "官浩杰"),
    ("lihongqi", "李红旗"), ("liwenbo", "李文波"), ("chenyu", "陈愉"),
    ("mojunyou", "莫钧友"), ("linyuping", "林宇萍"), ("wuyutian", "吴宇天"),
    ("caoyuan", "曹原"), ("liuxin01", "刘昕"),
]
WEEKLY_NAME_MAP = dict(WEEKLY_ROSTER, **{"liuying": "刘颖"})
WEEKLY_FILE_KEYS = {
    "new_tasks": "本周新增任务", "completed_tasks": "本周已完成任务",
    "new_defects": "本周新增缺陷", "fixed_defects": "本周已修复缺陷",
    "pending_tasks": "待处理任务", "pending_defects": "总挂起缺陷数",
}
WEEKLY_STAT_METRICS = (
    "本周新增任务数", "本周完成任务数", "本周新增缺陷数", "本周已修复缺陷数",
    "本周延期缺陷数", "总挂起缺陷数", "本周延期任务数", "总挂起任务数",
)

FIELD_ALIASES = {
    "issue_key": ["问题关键字", "问题键", "issuekey", "key"],
    "issue_id": ["问题id", "issueid"],
    "parent_id": ["父级id", "父问题id", "parentid"],
    "status": ["状态", "status"],
    "created_at": ["创建日期", "创建时间", "created", "createddate"],
    "summary": ["概要", "标题", "summary"],
    "assignee": ["经办人", "负责人", "assignee"],
    "reporter": ["报告人", "reporter"],
    "delivery_date": ["自定义字段交付日期", "交付日期", "deliverydate"],
    "resolution": ["解决结果", "resolution"],
    "issue_type": ["问题类型", "类型", "issuetype"],
    "developer": ["自定义字段开发人员", "开发人员", "developer"],
    "planned_completion": ["自定义字段计划完成日期", "计划完成日期", "plannedcompletion"],
}


def _notify(callback: ProgressCallback, stage: str, percent: int, detail: str) -> None:
    if callback:
        callback({"stage": stage, "percent": max(0, min(100, percent)), "detail": detail})


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_\-（）()：:]+", "", text)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M") if value.time() != datetime.min.time() else value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            converted = datetime(1899, 12, 30) + timedelta(days=float(value))
            return converted.strftime("%Y-%m-%d %H:%M") if converted.time() != datetime.min.time() else converted.strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            return _text(value)
    return _text(value)


def _excel_safe(value: Any) -> Any:
    """避免把用户输入中的 =、+、-、@ 当作 Excel 公式执行。"""
    if not isinstance(value, str):
        return value
    return "'" + value if value[:1] in {"=", "+", "-", "@"} else value


def _column_name(index: int) -> str:
    return get_column_letter(index + 1)


def _find_index(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [_norm_header(header) for header in headers]
    candidates = {_norm_header(alias) for alias in aliases}
    for index, value in enumerate(normalized):
        if value in candidates:
            return index
    return None


def _module_indexes(headers: list[str]) -> list[int]:
    indexes = []
    for index, header in enumerate(headers):
        if _norm_header(header).startswith("模块"):
            indexes.append(index)
    return indexes


def _field_mapping(headers: list[str], indexes: dict[str, int | None], module_indexes: list[int]) -> dict[str, dict[str, Any]]:
    mapping: dict[str, dict[str, Any]] = {}
    for field, index in indexes.items():
        mapping[field] = {
            "header": headers[index] if index is not None and index < len(headers) else "",
            "column": _column_name(index) if index is not None else "",
            "found": index is not None,
        }
    mapping["modules"] = {
        "header": "、".join(headers[index] for index in module_indexes),
        "column": ", ".join(_column_name(index) for index in module_indexes),
        "found": bool(module_indexes),
    }
    return mapping


def process_jira_workbook(path: str | Path, progress_callback: ProgressCallback = None) -> dict[str, Any]:
    """读取 Jira xlsx，并返回页面预览与导出共用的 JSON 结果。"""
    _notify(progress_callback, "读取工作表", 8, "正在打开 Excel 工作簿")
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = next((sheet for sheet in workbook.worksheets if sheet.max_row), None)
        if worksheet is None:
            raise ValueError("工作簿中没有可读取的数据")

        rows = worksheet.iter_rows()
        header_cells = next(rows, None)
        if not header_cells:
            raise ValueError("工作表为空")
        populated_header_indexes = [index for index, cell in enumerate(header_cells) if _text(cell.value)]
        if not populated_header_indexes:
            raise ValueError("首行没有可识别的字段名")
        header_cells = header_cells[: populated_header_indexes[-1] + 1]
        headers = [_text(cell.value) or f"列{index + 1}" for index, cell in enumerate(header_cells)]
        # 同名 Jira 字段保留全部列，避免第二个“模块”覆盖第一个。
        seen: Counter[str] = Counter()
        unique_headers = []
        for header in headers:
            seen[header] += 1
            unique_headers.append(header if seen[header] == 1 else f"{header}_{seen[header]}")

        indexes = {field: _find_index(headers, aliases) for field, aliases in FIELD_ALIASES.items()}
        module_indexes = _module_indexes(headers)
        missing = [field for field in ("status", "summary", "assignee") if indexes[field] is None]
        if missing:
            labels = {"status": "状态", "summary": "概要", "assignee": "经办人"}
            raise ValueError("缺少必需字段：" + "、".join(labels[field] for field in missing))

        mapping = _field_mapping(unique_headers, indexes, module_indexes)
        _notify(progress_callback, "识别字段", 18, f"已识别 {sum(item['found'] for item in mapping.values())} 个字段")

        records: list[dict[str, Any]] = []
        raw_rows: list[list[Any]] = []
        for row_number, cells in enumerate(rows, start=2):
            values = [cell.value for cell in cells]
            if not any(value not in (None, "") for value in values):
                continue
            raw_rows.append([_date_text(value) if isinstance(value, (date, datetime)) else _text(value) for value in values])
            get = lambda field: values[indexes[field]] if indexes[field] is not None and indexes[field] < len(values) else None
            issue_key = _text(get("issue_key")) or f"第{row_number}行"
            status_raw = _text(get("status"))
            summary = _text(get("summary"))
            assignee = _text(get("assignee"))
            developer = _text(get("developer"))
            reporter = _text(get("reporter"))
            if developer:
                summary_person = developer
                summary_source = "开发人员"
                summary_note = ""
            elif assignee:
                summary_person = assignee
                summary_source = "经办人"
                summary_note = "开发人员为空，按经办人汇总"
            elif reporter:
                summary_person = reporter
                summary_source = "报告人"
                summary_note = "开发人员和经办人均为空，按报告人汇总"
            else:
                summary_person = "未分配"
                summary_source = "未分配"
                summary_note = "开发人员、经办人和报告人均为空，归入未分配"
            modules = [_text(values[index]) for index in module_indexes if index < len(values) and _text(values[index])]
            consistency = "待补开发人员" if not developer else "一致" if assignee == developer else "人员不一致"
            display_status = STATUS_MAP.get(status_raw, status_raw or "未填写")
            # 汇总归属依次使用开发人员、经办人、报告人；均为空时归入未分配。
            # 人员字段不一致只做异常提示，不影响按上述优先级归入汇总。
            record = {
                "source_row": row_number,
                "issue_key": issue_key,
                "issue_id": _text(get("issue_id")),
                "parent_id": _text(get("parent_id")),
                "status_raw": status_raw,
                "status": display_status,
                "created_at": _date_text(get("created_at")),
                "summary": summary,
                "assignee": assignee or "未分配",
                "reporter": reporter,
                "developer": developer,
                "summary_person": summary_person,
                "summary_source": summary_source,
                "summary_note": summary_note,
                "modules": modules,
                "module": "、".join(dict.fromkeys(modules)),
                "delivery_date": _date_text(get("delivery_date")),
                "resolution": _text(get("resolution")),
                "issue_type": _text(get("issue_type")),
                "planned_completion": _date_text(get("planned_completion")),
                "consistency": consistency,
                "included": True,
                "summary_text": f"{summary}【{display_status}】",
            }
            records.append(record)

        _notify(progress_callback, "清洗数据", 42, f"已读取 {len(records)} 条任务")
        key_counts = Counter(record["issue_key"] for record in records)
        roster_order = {name: index + 1 for index, name in enumerate(DEFAULT_ROSTER)}
        for record in records:
            record["sort_order"] = roster_order.get(record["summary_person"], 9999)
            record["duplicate"] = key_counts[record["issue_key"]] > 1

        def sort_key(record: dict[str, Any]) -> tuple[Any, ...]:
            return (
                record["sort_order"],
                record["planned_completion"] or "9999-99-99",
                record["delivery_date"] or "9999-99-99",
                record["source_row"],
            )

        records.sort(key=sort_key)
        included = [record for record in records if record["included"]]
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in included:
            grouped[record["summary_person"]].append(record)

        summaries = []
        summary_names = DEFAULT_ROSTER + sorted(name for name in grouped if name not in roster_order)
        for name in summary_names:
            items = grouped.get(name, [])
            status_count = Counter(item["status"] for item in items)
            summaries.append({
                "name": name,
                "task_count": len(items),
                "completed_count": status_count.get("已完成", 0),
                "in_progress_count": status_count.get("进行中", 0),
                "open_count": status_count.get("开放", 0),
                "merged_text": ";\n".join(item["summary_text"] for item in items),
                "items": [item["issue_key"] for item in items],
            })

        anomalies = []
        for record in records:
            if record["consistency"] == "待补开发人员":
                anomalies.append({"row": record["source_row"], "issue_key": record["issue_key"], "type": "developer_missing", "label": "待补开发人员", "detail": record["summary_note"]})
            elif record["consistency"] == "人员不一致":
                anomalies.append({"row": record["source_row"], "issue_key": record["issue_key"], "type": "developer_mismatch", "label": "人员不一致", "detail": f"经办人：{record['assignee']}；开发人员：{record['developer']}；已按开发人员纳入人员汇总"})
            if record["status_raw"] not in KNOWN_STATUSES:
                anomalies.append({"row": record["source_row"], "issue_key": record["issue_key"], "type": "unknown_status", "label": "未知状态", "detail": record["status_raw"] or "状态为空"})
            if not record["planned_completion"]:
                anomalies.append({"row": record["source_row"], "issue_key": record["issue_key"], "type": "planned_date_missing", "label": "计划完成日期缺失", "detail": "原始字段为空"})
            if record["duplicate"]:
                anomalies.append({"row": record["source_row"], "issue_key": record["issue_key"], "type": "duplicate_issue", "label": "问题关键字重复", "detail": "建议回到 Jira 导出数据确认是否重复"})

        status_counts = Counter(record["status_raw"] for record in records)
        _notify(progress_callback, "核验并分组", 78, f"按人员归属汇总 {len(included)} 条，异常 {len(anomalies)} 条")
        result = {
            "ok": True,
            "source": Path(path).name,
            "sheet": worksheet.title,
            "field_mapping": mapping,
            "stats": {
                "total_rows": len(records),
                "included_rows": len(included),
                "pending_developer": sum(record["consistency"] == "待补开发人员" for record in records),
                "mismatched_developer": sum(record["consistency"] == "人员不一致" for record in records),
                "unique_assignees": len({record["summary_person"] for record in records}),
                "duplicate_keys": sum(count > 1 for count in key_counts.values()),
                "missing_planned_completion": sum(not record["planned_completion"] for record in records),
                "status_counts": dict(status_counts),
            },
            "summaries": summaries,
            "records": records,
            "anomalies": anomalies,
            "raw_headers": unique_headers,
            "raw_rows": raw_rows,
        }
        _notify(progress_callback, "生成结果", 100, "Jira 数据处理完成")
        return result
    finally:
        workbook.close()


def process_daily_jira_workbook(path: str | Path, progress_callback: ProgressCallback = None) -> dict[str, Any]:
    """按开发人员、经办人、报告人的优先级生成每日 Jira 汇总。"""
    _notify(progress_callback, "读取工作表", 8, "正在打开 Excel 工作簿")
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = next((sheet for sheet in workbook.worksheets if sheet.max_row), None)
        if worksheet is None:
            raise ValueError("工作簿中没有可读取的数据")

        rows = worksheet.iter_rows()
        header_cells = next(rows, None)
        if not header_cells:
            raise ValueError("工作表为空")
        populated_header_indexes = [index for index, cell in enumerate(header_cells) if _text(cell.value)]
        if not populated_header_indexes:
            raise ValueError("首行没有可识别的字段名")
        header_cells = header_cells[: populated_header_indexes[-1] + 1]
        headers = [_text(cell.value) or f"列{index + 1}" for index, cell in enumerate(header_cells)]
        seen: Counter[str] = Counter()
        unique_headers = []
        for header in headers:
            seen[header] += 1
            unique_headers.append(header if seen[header] == 1 else f"{header}_{seen[header]}")

        indexes = {field: _find_index(headers, aliases) for field, aliases in FIELD_ALIASES.items()}
        missing = [field for field in ("status", "summary") if indexes[field] is None]
        if missing:
            labels = {"status": "状态", "summary": "概要"}
            raise ValueError("缺少每日处理必需字段：" + "、".join(labels[field] for field in missing))
        if all(indexes[field] is None for field in ("developer", "assignee", "reporter")):
            raise ValueError("缺少每日处理人员字段：开发人员、经办人或报告人")

        mapping = _field_mapping(unique_headers, indexes, _module_indexes(headers))
        _notify(progress_callback, "识别字段", 18, "已识别状态、概要和人员归属字段")

        records: list[dict[str, Any]] = []
        for row_number, cells in enumerate(rows, start=2):
            values = [cell.value for cell in cells]
            if not any(value not in (None, "") for value in values):
                continue
            get = lambda field: values[indexes[field]] if indexes[field] is not None and indexes[field] < len(values) else None
            assignee = _text(get("assignee"))
            developer = _text(get("developer"))
            reporter = _text(get("reporter"))
            status_raw = _text(get("status"))
            status = STATUS_MAP.get(status_raw, status_raw or "未填写")
            summary = _text(get("summary"))
            if developer:
                summary_person = developer
                summary_source = "开发人员"
                consistency = "相同" if not assignee or assignee == developer else "不同"
                summary_note = "" if consistency == "相同" else "经办人与开发人员不一致，按开发人员汇总"
            elif assignee:
                summary_person = assignee
                summary_source = "经办人"
                consistency = "缺少开发人员"
                summary_note = "开发人员为空，按经办人汇总"
            elif reporter:
                summary_person = reporter
                summary_source = "报告人"
                consistency = "缺少开发人员和经办人"
                summary_note = "开发人员和经办人均为空，按报告人汇总"
            else:
                summary_person = "未分配"
                summary_source = "未分配"
                consistency = "人员缺失"
                summary_note = "开发人员、经办人和报告人均为空，未纳入每日汇总"
            included = bool(developer or assignee or reporter)
            records.append({
                "source_row": row_number,
                "issue_key": _text(get("issue_key")) or f"第{row_number}行",
                "status_raw": status_raw,
                "status": status,
                "summary": summary,
                "assignee": assignee,
                "reporter": reporter,
                "developer": developer,
                "summary_person": summary_person,
                "summary_source": summary_source,
                "summary_note": summary_note,
                "module": "",
                "planned_completion": _date_text(get("planned_completion")),
                "consistency": consistency,
                "included": included,
                "summary_text": f"{summary}【{status}】",
            })

        _notify(progress_callback, "确定人员归属", 48, f"已读取 {len(records)} 条任务")
        roster_order = {name: index + 1 for index, name in enumerate(DEFAULT_ROSTER)}
        records.sort(key=lambda record: (
            0 if record["included"] else 1,
            roster_order.get(record["summary_person"], len(DEFAULT_ROSTER) + 1),
            record["source_row"],
        ))
        included_records = [record for record in records if record["included"]]
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for record in included_records:
            grouped[record["summary_person"]].append(record)

        summary_names = DEFAULT_ROSTER + sorted(name for name in grouped if name not in roster_order)
        summaries = []
        for name in summary_names:
            items = grouped.get(name, [])
            status_count = Counter(item["status"] for item in items)
            summaries.append({
                "name": name,
                "task_count": len(items),
                "completed_count": status_count.get("已完成", 0),
                "in_progress_count": status_count.get("进行中", 0),
                "open_count": status_count.get("开放", 0),
                "merged_text": ";\n".join(item["summary_text"] for item in items),
                "items": [item["issue_key"] for item in items],
            })

        anomalies = [{
            "row": record["source_row"],
            "issue_key": record["issue_key"],
            "type": "daily_person_fallback" if record["included"] else "daily_person_missing",
            "label": record["consistency"],
            "detail": record["summary_note"],
        } for record in records if record["summary_note"]]
        _notify(progress_callback, "排序并合并", 82, f"已按人员优先级纳入 {len(included_records)} 条记录，按 {len(summaries)} 人生成汇总")
        result = {
            "ok": True,
            "mode": "daily",
            "source": Path(path).name,
            "sheet": worksheet.title,
            "field_mapping": mapping,
            "stats": {
                "total_rows": len(records),
                "included_rows": len(included_records),
                "excluded_rows": len(records) - len(included_records),
                "pending_developer": sum(not record["developer"] for record in records),
                "mismatched_developer": sum(record["consistency"] == "不同" for record in records),
                "unique_assignees": len(grouped),
                "status_counts": dict(Counter(record["status_raw"] for record in records)),
            },
            "summaries": summaries,
            "records": records,
            "anomalies": anomalies,
        }
        _notify(progress_callback, "生成结果", 100, "每日 Jira 数据处理完成")
        return result
    finally:
        workbook.close()


def discover_weekly_files(paths: list[str | Path]) -> dict[str, Path]:
    """按 Jira 导出文件名识别周报六类来源，忽略 Excel 临时锁定文件。"""
    matches: dict[str, list[Path]] = {key: [] for key in WEEKLY_FILE_KEYS}
    for raw_path in paths:
        path = Path(raw_path)
        if path.name.startswith("~$") or path.suffix.lower() != ".xlsx":
            continue
        for key, marker in WEEKLY_FILE_KEYS.items():
            if marker in path.name:
                matches[key].append(path)
                break
    conflicts = {key: items for key, items in matches.items() if len(items) > 1}
    if conflicts:
        detail = "；".join(f"{key}: {', '.join(item.name for item in items)}" for key, items in conflicts.items())
        raise ValueError(f"周报文件存在多个候选文件：{detail}")
    missing = [key for key, items in matches.items() if not items]
    if missing:
        labels = "、".join(WEEKLY_FILE_KEYS[key] for key in missing)
        raise ValueError(f"缺少周报来源文件：{labels}")
    return {key: items[0] for key, items in matches.items()}


def _read_weekly_records(path: str | Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = next((sheet for sheet in workbook.worksheets if sheet.max_row), None)
        if worksheet is None:
            raise ValueError(f"工作簿为空：{Path(path).name}")
        rows = worksheet.iter_rows()
        header_cells = next(rows, None)
        if not header_cells:
            raise ValueError(f"工作表为空：{Path(path).name}")
        last = max((index for index, cell in enumerate(header_cells) if _text(cell.value)), default=-1)
        if last < 0:
            raise ValueError(f"首行没有字段名：{Path(path).name}")
        headers = [_text(cell.value) or f"列{index + 1}" for index, cell in enumerate(header_cells[:last + 1])]
        indexes = {field: _find_index(headers, aliases) for field, aliases in FIELD_ALIASES.items()}
        missing = [field for field in ("issue_key", "status", "summary") if indexes[field] is None]
        if missing:
            labels = {"issue_key": "问题关键字", "status": "状态", "summary": "概要"}
            raise ValueError(f"{Path(path).name} 缺少必需字段：{'、'.join(labels[field] for field in missing)}")
        module_indexes = _module_indexes(headers)
        records = []
        for row_number, cells in enumerate(rows, start=2):
            values = [cell.value for cell in cells]
            if not any(value not in (None, "") for value in values):
                continue
            get = lambda field: values[indexes[field]] if indexes[field] is not None and indexes[field] < len(values) else None
            status_raw = _text(get("status"))
            resolution = _text(get("resolution"))
            issue_type = _text(get("issue_type"))
            assignee = _text(get("assignee"))
            developer = _text(get("developer"))
            reporter = _text(get("reporter"))
            records.append({
                "source_row": row_number, "issue_key": _text(get("issue_key")) or f"第{row_number}行",
                "issue_id": _text(get("issue_id")), "status": status_raw, "summary": _text(get("summary")),
                "created_at": get("created_at"), "assignee": assignee, "reporter": reporter,
                "developer": developer, "module": "、".join(dict.fromkeys(_text(values[i]) for i in module_indexes if i < len(values) and _text(values[i]))),
                "delivery_date": get("delivery_date"), "resolution": resolution, "issue_type": issue_type,
                "planned_completion": get("planned_completion"),
            })
        return records, {"file": Path(path).name, "sheet": worksheet.title, "headers": headers}
    finally:
        workbook.close()


def preview_jira_workbook(path: str | Path) -> dict[str, Any]:
    """Return a compact, safe preview of an uploaded Jira workbook."""
    records, metadata = _read_weekly_records(path)
    columns = ["问题关键字", "状态", "概要", "经办人", "开发人员", "问题类型"]
    fields = ["issue_key", "status", "summary", "assignee", "developer", "issue_type"]
    rows = [[_weekly_display_person(record.get(field, "")) if field in {"assignee", "developer"} else _text(record.get(field, "")) for field in fields] for record in records[:15]]
    return {"file": metadata["file"], "sheet": metadata["sheet"], "headers": columns, "rows": rows, "total_rows": len(records)}


def _is_defect(record: dict[str, Any]) -> bool:
    return _text(record.get("issue_type")).lower() in {"故障", "缺陷", "bug", "defect"}


def _is_completed(record: dict[str, Any]) -> bool:
    status = _text(record.get("status"))
    resolution = _text(record.get("resolution"))
    return bool(resolution) and (status in {"已解决", "已关闭", "已完成"} or resolution in {"已解决", "已完成", "完成"})


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _weekly_owners(record: dict[str, Any]) -> tuple[list[str], str]:
    assignee = _text(record.get("assignee"))
    developer = _text(record.get("developer"))
    reporter = _text(record.get("reporter"))
    if assignee and developer and assignee != developer:
        return list(dict.fromkeys([assignee, developer])), "经办人与开发人员不一致，按经办人基础统计并追加开发人员"
    if assignee:
        return [assignee], "开发人员为空，按经办人统计" if not developer else "经办人与开发人员一致"
    if developer:
        return [developer], "经办人为空，按开发人员统计"
    if reporter:
        return [reporter], "经办人和开发人员为空，按报告人兜底"
    return ["未分配"], "经办人、开发人员和报告人均为空"


def _weekly_display_person(value: Any) -> str:
    """将 Jira 用户名转换为周报明细中的中文姓名。"""
    text = _text(value)
    return WEEKLY_NAME_MAP.get(text, text)


def process_weekly_statistics(sources: dict[str, str | Path], progress_callback: ProgressCallback = None) -> dict[str, Any]:
    """处理当前周六类 Jira 导出，生成 23 人周报统计及明细。"""
    all_records: dict[str, list[dict[str, Any]]] = {}
    metadata = {}
    for index, (key, path) in enumerate(sources.items(), start=1):
        _notify(progress_callback, "读取周报文件", index * 10, f"正在读取 {Path(path).name}")
        all_records[key], metadata[key] = _read_weekly_records(path)
    # 已完成任务导出通常不带人员和问题类型，按问题关键字回填新增任务中的归属字段。
    new_task_lookup = {record["issue_key"]: record for record in all_records["new_tasks"]}
    for record in all_records["completed_tasks"]:
        reference = new_task_lookup.get(record["issue_key"])
        if reference:
            for field in ("assignee", "developer", "reporter", "issue_type", "delivery_date", "module"):
                if not record.get(field):
                    record[field] = reference.get(field)
    created_dates = [_date_value(item.get("created_at")) for rows in all_records.values() for item in rows]
    report_date = max((value for value in created_dates if value), default=date.today())
    roster = {username: {"序号": index, "姓名": name, "username": username} for index, (username, name) in enumerate(WEEKLY_ROSTER, start=1)}
    summary = [{**person, **{key: 0 for key in ("new_task_count", "completed_task_count", "new_defect_count", "fixed_defect_count", "delayed_defect_count", "pending_defect_count", "delayed_task_count", "pending_task_count")}} for person in roster.values()]
    by_username = {item["username"]: item for item in summary}
    anomalies: list[dict[str, Any]] = []
    sections = {"delayed_defects": [], "pending_defects": [], "delayed_tasks": [], "pending_tasks": []}

    def add_count(record: dict[str, Any], field: str, reason: str) -> None:
        owners, note = _weekly_owners(record)
        record = dict(record, owner_note=note, source_field=field)
        for owner in owners:
            target = by_username.get(owner)
            if target:
                target[field] += 1
            else:
                anomalies.append({"source": field, "issue_key": record["issue_key"], "label": "未知人员", "detail": owner})
        if note != "经办人与开发人员一致":
            anomalies.append({"source": field, "issue_key": record["issue_key"], "label": "人员归属备注", "detail": note})
        if not record.get("resolution"):
            anomalies.append({"source": field, "issue_key": record["issue_key"], "label": "解决结果为空", "detail": "未计入完成或修复统计"})
        return record

    for record in all_records["new_tasks"]:
        if not _is_defect(record):
            add_count(record, "new_task_count", "新增任务")
    for record in all_records["completed_tasks"]:
        if not _is_defect(record) and _is_completed(record):
            add_count(record, "completed_task_count", "完成任务")
    for record in all_records["new_defects"]:
        if _is_defect(record):
            add_count(record, "new_defect_count", "新增缺陷")
    for record in all_records["fixed_defects"]:
        if _is_defect(record) and _is_completed(record):
            add_count(record, "fixed_defect_count", "修复缺陷")

    def classify_pending(key: str, defect_field: str, task_field: str, section_key: str, require_overdue: bool = False) -> None:
        for record in all_records[key]:
            defect = _is_defect(record)
            if require_overdue:
                delivery = _date_value(record.get("delivery_date"))
                if not delivery:
                    anomalies.append({"source": key, "issue_key": record["issue_key"], "label": "交付日期为空", "detail": "无法判断逾期"})
                    continue
                if delivery >= report_date or _is_completed(record):
                    continue
            elif _is_completed(record):
                continue
            if (defect and defect_field) or (not defect and task_field):
                field = defect_field if defect else task_field
                detail = add_count(record, field, section_key)
                sections[section_key].append(detail)

    classify_pending("new_defects", "delayed_defect_count", "", "delayed_defects", True)
    classify_pending("pending_defects", "pending_defect_count", "", "pending_defects")
    classify_pending("pending_tasks", "", "delayed_task_count", "delayed_tasks", True)
    classify_pending("pending_tasks", "", "pending_task_count", "pending_tasks")
    return {"ok": True, "report_date": report_date.isoformat(), "sources": metadata, "summary": summary, "sections": sections, "anomalies": anomalies, "stats": {"total_rows": sum(len(rows) for rows in all_records.values()), "anomaly_count": len(anomalies)}}


def _extract_screenshot_totals(path: str | Path) -> dict[str, int]:
    """Read visible Chinese names and the value in the screenshot's 合计 column."""
    ocr = RapidOCR()
    ocr.text_score = 0.2
    ocr.min_height = 3
    detected, _ = ocr(str(path))
    rows: dict[int, dict[str, Any]] = defaultdict(lambda: {"names": [], "numbers": []})
    total_column_centers: list[float] = []
    roster_names = {name for _, name in WEEKLY_ROSTER}
    header_people: list[tuple[float, str]] = []
    summary_total_row: int | None = None
    for item in detected or []:
        if len(item) < 2:
            continue
        box, text = item[0], _text(item[1])
        if not text:
            continue
        center_x = sum(point[0] for point in box) / 4
        center_y = round(sum(point[1] for point in box) / 4 / 12) * 12
        if "合计" in text:
            total_column_centers.append(center_x)
        if text in roster_names:
            header_people.append((center_x, text))
        if "唯一问题合计" in text:
            summary_total_row = center_y
        row = rows[center_y]
        numbers = [int(value) for value in re.findall(r"\d+", text)]
        if numbers:
            row["numbers"].extend((center_x, number) for number in numbers)
        name = re.sub(r"[：:，,。.!！?？\s]", "", text)
        if re.search(r"[\u4e00-\u9fff]", name) and not any(token in name for token in ("经办人", "合计", "唯一问题", "开放", "处理中", "统计", "分组")):
            row["names"].append(name)
    if header_people and summary_total_row is not None:
        # Some Jira pivots put people in the column headers and issue types in rows.
        # In that layout the bottom "唯一问题合计" row is the per-person result;
        # the grey rightmost 合计 column is only a cross-check and must be excluded.
        total_x = max(total_column_centers, default=float("inf"))
        people_right_edge = max(center for center, _ in header_people)
        boundary = (people_right_edge + total_x) / 2 if total_x != float("inf") else float("inf")
        column_totals: dict[str, int] = {}
        for center_x, number in rows[summary_total_row]["numbers"]:
            if center_x >= boundary:
                continue
            _, person = min(header_people, key=lambda item: abs(item[0] - center_x))
            column_totals[person] = number
        if column_totals:
            return column_totals
    if total_column_centers:
        # The total cell can be faint against the table grid. Re-read the total column at
        # higher resolution and mark those values explicitly, so other numeric columns
        # cannot be mistaken for totals when the first OCR pass misses a cell. Reuse one
        # OCR session for every cell: constructing RapidOCR for each roster row repeatedly
        # loads an ONNX session and can make a normal screenshot exceed the UI timeout.
        image = cv2.imread(str(path))
        if image is not None:
            cell_ocr = RapidOCR()
            cell_ocr.text_score = 0.05
            cell_ocr.min_height = 1
            total_x = max(total_column_centers)
            left = max(0, int(total_x - 70))
            right = min(image.shape[1], int(total_x + 100))
            for row_y, row in rows.items():
                if not row["names"]:
                    continue
                top = max(0, row_y - 20)
                bottom = min(image.shape[0], row_y + 20)
                cell_image = cv2.cvtColor(image[top:bottom, left:right], cv2.COLOR_BGR2GRAY)
                cell_image = cv2.resize(cell_image, None, fx=5, fy=5, interpolation=cv2.INTER_CUBIC)
                cell_detected, _ = cell_ocr(cell_image)
                values = [int(value) for item in cell_detected or [] for value in re.findall(r"\d+", _text(item[1]))]
                if values:
                    row["total_numbers"] = values[-1]
    totals = {}
    for row in rows.values():
        if not row["names"] or (not row["numbers"] and "total_numbers" not in row):
            continue
        if "total_numbers" in row:
            totals[row["names"][0]] = row["total_numbers"]
        elif total_column_centers:
            total_x = max(total_column_centers)
            totals[row["names"][0]] = min(row["numbers"], key=lambda item: abs(item[0] - total_x))[1]
        else:
            # 某些导出截图会裁掉表头；没有列定位信息时保留最右侧数值作为兼容回退。
            totals[row["names"][0]] = row["numbers"][-1][1]
    return totals


def _validate_screenshot_metric(path: str | Path, expected_metric: str) -> None:
    """Ensure the visible Jira pivot title belongs to the selected statistic row."""
    # A RapidOCR instance owns an ONNX runtime session. Reusing one global
    # instance across concurrent browser checks can leave later jobs waiting
    # forever at the screenshot-recognition stage.
    ocr = RapidOCR()
    ocr.text_score = 0.2
    ocr.min_height = 3
    image = cv2.imread(str(path))
    if image is None:
        raise ValueError("无法读取截图，请重新粘贴")
    # The Jira statistic title is always in the header. Avoid a full-page OCR pass
    # while the user is waiting for immediate upload validation.
    header = image[:max(140, int(image.shape[0] * 0.18)), :]
    if header.shape[1] > 1400:
        scale = 1400 / header.shape[1]
        header = cv2.resize(header, None, fx=scale, fy=scale, interpolation=cv2.INTER_AREA)
    detected, _ = ocr(header)
    text = "".join(_text(item[1]) for item in detected or [] if len(item) >= 2)
    actual = next((metric for metric in WEEKLY_STAT_METRICS if metric in text), None)
    if actual is None:
        # Browser chrome and custom Jira layouts can push the title below the fast crop.
        # Fall back to one full-image pass before rejecting a valid screenshot.
        detected, _ = ocr(str(path))
        text = "".join(_text(item[1]) for item in detected or [] if len(item) >= 2)
        actual = next((metric for metric in WEEKLY_STAT_METRICS if metric in text), None)
    if actual is None:
        raise ValueError(f"截图未识别到统计标题，无法确认是否为“{expected_metric}”，请重新截图")
    if actual != expected_metric:
        raise ValueError(f"截图统计标题为“{actual}”，与当前“{expected_metric}”不一致，不能生成文件")


def validate_screenshot_metric(path: str | Path, expected_metric: str) -> None:
    """Validate a screenshot title before it is accepted in the browser."""
    _validate_screenshot_metric(path, expected_metric)


def _extract_screenshot_grand_total(path: str | Path) -> int | None:
    """Read the grey 合计 value from the Jira pivot's 唯一问题合计 row."""
    ocr = RapidOCR()
    ocr.text_score = 0.2
    ocr.min_height = 3
    detected, _ = ocr(str(path))
    total_x: float | None = None
    summary_y: int | None = None
    values: list[tuple[float, int]] = []
    for item in detected or []:
        if len(item) < 2:
            continue
        box, text = item[0], _text(item[1])
        if not text:
            continue
        center_x = sum(point[0] for point in box) / 4
        center_y = round(sum(point[1] for point in box) / 4 / 12) * 12
        if "合计" in text and "唯一问题" not in text:
            total_x = max(total_x or center_x, center_x)
        if "唯一问题合计" in text:
            summary_y = center_y
        values.extend((center_x, int(value)) for value in re.findall(r"\d+", text) if center_y == summary_y)
    if total_x is None or summary_y is None:
        return None
    summary_values = []
    for item in detected or []:
        if len(item) < 2:
            continue
        box, text = item[0], _text(item[1])
        center_y = round(sum(point[1] for point in box) / 4 / 12) * 12
        if center_y == summary_y:
            center_x = sum(point[0] for point in box) / 4
            summary_values.extend((center_x, int(value)) for value in re.findall(r"\d+", text))
    return min(summary_values, key=lambda item: abs(item[0] - total_x))[1] if summary_values else None


def process_new_task_statistics(path: str | Path, screenshot: str | Path, metric: str = "本周新增任务数", progress_callback: ProgressCallback = None) -> dict[str, Any]:
    """Build new-task counts from the Jira export and correct owner mismatches.

    The screenshot is a total check, not a reliable row-level data source: Jira
    pivot screenshots contain small, grid-separated names and totals that OCR can
    partially miss.  The accompanying Excel has the complete assignee column, so
    it is the source of truth for every person's baseline count.
    """
    _notify(progress_callback, "校验截图", 25, "正在确认统计标题与当前指标一致")
    _validate_screenshot_metric(screenshot, metric)
    _notify(progress_callback, "读取 Jira 数据", 45, "正在检查开发人员和经办人")
    records, metadata = _read_weekly_records(path)
    _notify(progress_callback, "核对截图合计", 65, "正在核对截图总数与 Jira Excel")
    screenshot_total = _extract_screenshot_grand_total(screenshot)
    excel_total = sum(not _is_defect(record) for record in records)
    if screenshot_total is not None and screenshot_total != excel_total:
        raise ValueError(f"Jira Excel 任务数为 {excel_total}，截图合计为 {screenshot_total}，数据不一致，请确认上传的是同一批周报文件和截图")
    counts = Counter(
        _text(record.get("assignee"))
        for record in records
        if not _is_defect(record) and _text(record.get("assignee"))
    )
    mismatches = []
    for record in records:
        if _is_defect(record):
            continue
        assignee = _text(record.get("assignee"))
        developer = _text(record.get("developer"))
        if assignee and developer and assignee != developer:
            counts[developer] += 1
            counts[assignee] = max(0, counts[assignee] - 1)
            mismatches.append({"issue_key": record["issue_key"], "assignee": _weekly_display_person(assignee), "developer": _weekly_display_person(developer)})
    summary = [{"序号": index, "姓名": name, metric: counts.get(username) or ""} for index, (username, name) in enumerate(WEEKLY_ROSTER, start=1)]
    _notify(progress_callback, "修正人员归属", 80, "已按开发人员调整任务数量")
    return {"ok": True, "source": metadata, "summary": summary, "mismatches": mismatches, "screenshot_total": screenshot_total}


def process_screenshot_statistics(screenshot: str | Path, metric: str, progress_callback: ProgressCallback = None) -> dict[str, Any]:
    """Extract one Jira screenshot statistic in the fixed weekly roster order."""
    _notify(progress_callback, "校验截图", 30, "正在确认统计标题与当前指标一致")
    _validate_screenshot_metric(screenshot, metric)
    _notify(progress_callback, "识别统计截图", 65, "正在提取人员和合计数量")
    baseline = _extract_screenshot_totals(screenshot)
    summary = [{"序号": index, "姓名": name, metric: baseline.get(name) or ""} for index, (_, name) in enumerate(WEEKLY_ROSTER, start=1)]
    return {"ok": True, "summary": summary, "screenshot_totals": baseline, "metric": metric}


def export_new_task_statistics_xlsx(result: dict[str, Any], target: str | Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "本周新增任务"
    metric = next((key for key in result["summary"][0] if key not in {"序号", "姓名"}), "本周新增任务数")
    sheet.append(["姓名", metric])
    header_fill = PatternFill("solid", fgColor="2F5597")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for item in result["summary"]:
        sheet.append([item["姓名"], item.get(metric, "")])
    total_row = sheet.max_row + 1
    sheet.cell(total_row, 1, "合计")
    sheet.cell(total_row, 2, f"=SUM(B2:B{total_row - 1})")
    for cell in sheet[total_row]:
        cell.font = Font(bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(left=Side(style="thin", color="808080"), right=Side(style="thin", color="808080"), top=Side(style="thin", color="808080"), bottom=Side(style="thin", color="808080"))
            cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 22
    sheet.freeze_panes = "A2"
    workbook.save(target)
    workbook.close()


def export_combined_weekly_statistics_xlsx(statistics: dict[str, list[dict[str, Any]]], target: str | Path) -> None:
    """Export the eight independently generated Jira statistics into one weekly sheet."""
    metrics = [
        ("new-tasks", "本周新增任务数"), ("completed-tasks", "本周完成任务数"),
        ("new-defects", "本周新增缺陷数"), ("fixed-defects", "本周已修复缺陷数"),
        ("delayed-defects", "本周延期缺陷数"), ("pending-defects", "总挂起缺陷数"),
        ("delayed-tasks", "本周延期任务数"), ("pending-tasks", "总挂起任务数"),
    ]
    values_by_metric = {
        key: {str(item.get("姓名", "")): item.get(label, "") for item in statistics.get(key, [])}
        for key, label in metrics
    }
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "周报统计汇总"
    sheet.append(["序号", "姓名", *(label for _, label in metrics)])
    fill = PatternFill("solid", fgColor="2F5597")
    thin = Side(style="thin", color="808080")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, (_, name) in enumerate(WEEKLY_ROSTER, start=1):
        row = [index, name]
        row.extend(values_by_metric[key].get(name, "") for key, _ in metrics)
        sheet.append(row)
    total_row = sheet.max_row + 1
    sheet.cell(total_row, 1, "合计")
    for column in range(3, 11):
        sheet.cell(total_row, column, f"=SUM({get_column_letter(column)}2:{get_column_letter(column)}{total_row - 1})")
    def is_positive(value: Any) -> bool:
        try:
            return float(value) > 0
        except (TypeError, ValueError):
            return False

    warning_columns = (7, 8, 9, 10)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(2, total_row):
        for column in warning_columns:
            if is_positive(sheet.cell(row, column).value):
                sheet.cell(row, column).font = Font(color="FF0000")
    for cell in sheet[total_row]:
        cell.font = Font(bold=True)
    for column in warning_columns:
        if any(is_positive(sheet.cell(row, column).value) for row in range(2, total_row)):
            sheet.cell(total_row, column).font = Font(color="FF0000", bold=True)
    for index, width in enumerate([8, 14, 16, 16, 16, 16, 16, 16, 16, 16], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    workbook.save(target)
    workbook.close()


def export_weekly_statistics_xlsx(result: dict[str, Any], target: str | Path) -> None:
    """导出当前周统计、延期挂起明细、异常和处理说明。"""
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "周报统计"
    headers = ["序号", "姓名", "本周新增任务数", "本周完成任务数", "本周新增缺陷数", "本周已修复缺陷数", "本周延期缺陷数", "总挂起缺陷数", "本周延期任务数", "总挂起任务数"]
    summary_sheet.append(headers)
    summary_sheet.row_dimensions[1].height = 30
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(color="FFFFFF", bold=True)
    red_font = Font(color="FF0000")
    for cell in summary_sheet[1]:
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fields = ["new_task_count", "completed_task_count", "new_defect_count", "fixed_defect_count", "delayed_defect_count", "pending_defect_count", "delayed_task_count", "pending_task_count"]
    for item in result["summary"]:
        values = [item["序号"], item["姓名"], item["new_task_count"], item["completed_task_count"], item["new_defect_count"], item["fixed_defect_count"]]
        values.extend(item[field] or None for field in fields[4:])
        summary_sheet.append(values)
        summary_sheet.row_dimensions[summary_sheet.max_row].height = 22
    total_row = summary_sheet.max_row + 1
    summary_sheet.cell(total_row, 1, "合计")
    summary_sheet.row_dimensions[total_row].height = 24
    for column in range(3, 11):
        summary_sheet.cell(total_row, column, f"=SUM({get_column_letter(column)}2:{get_column_letter(column)}{total_row - 1})")
    thin = Side(style="thin", color="808080")
    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in range(2, total_row):
        for column in (7, 8, 9, 10):
            if summary_sheet.cell(row, column).value:
                summary_sheet.cell(row, column).font = red_font
    for index, width in enumerate([8, 12, 16, 16, 16, 16, 16, 16, 16, 16], start=1):
        summary_sheet.column_dimensions[get_column_letter(index)].width = width
    summary_sheet.freeze_panes = "A2"
    summary_sheet.sheet_view.showGridLines = True
    summary_sheet.page_setup.orientation = "landscape"
    summary_sheet.page_setup.fitToWidth = 1
    summary_sheet.sheet_properties.pageSetUpPr.fitToPage = True

    detail_sheet = summary_sheet
    detail_headers = ["问题关键字", "状态", "创建日期", "概要", "经办人", "报告人", "模块", "交付日期", "解决结果", "问题类型", "开发人员", "计划完成日期"]
    detail_row = 1
    section_titles = [("delayed_defects", "本周延期缺陷数"), ("pending_defects", "总挂起缺陷数"), ("delayed_tasks", "本周延期任务数")]
    detail_row = total_row + 2
    for section_key, title in section_titles:
        detail_sheet.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=len(detail_headers))
        title_cell = detail_sheet.cell(detail_row, 1, title)
        title_cell.fill = PatternFill("solid", fgColor="D9E2F3")
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        detail_sheet.row_dimensions[detail_row].height = 28
        detail_row += 1
        for column, value in enumerate(detail_headers, start=1):
            detail_sheet.cell(detail_row, column, value)
        for cell in detail_sheet[detail_row]:
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(color="000000", bold=False)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        detail_sheet.row_dimensions[detail_row].height = 32
        detail_row += 1
        for record in result["sections"].get(section_key, []):
            values = [record.get(field) for field in ["issue_key", "status", "created_at", "summary", "assignee", "reporter", "module", "delivery_date", "resolution", "issue_type", "developer", "planned_completion"]]
            values[4] = _weekly_display_person(values[4])
            values[5] = _weekly_display_person(values[5])
            values[10] = _weekly_display_person(values[10])
            for column, value in enumerate(values, start=1):
                detail_sheet.cell(detail_row, column, value)
            summary_lines = max(1, (len(str(record.get("summary") or "")) + 28) // 29)
            detail_sheet.row_dimensions[detail_row].height = min(22 * summary_lines, 82)
            detail_row += 1
        detail_row += 1
    for row in detail_sheet.iter_rows(min_row=total_row + 2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row > total_row + 2 and cell.column in (3, 8, 12) and cell.value:
                cell.number_format = "yyyy/m/d h:mm"
    for section_offset in range(len(section_titles)):
        section_title_row = total_row + 2
        for previous_key, _ in section_titles[:section_offset]:
            section_title_row += 3 + len(result["sections"].get(previous_key, []))
        title_cell = detail_sheet.cell(section_title_row, 1)
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for index, width in enumerate([18, 12, 20, 42, 16, 16, 24, 20, 16, 14, 16, 20], start=1):
        detail_sheet.column_dimensions[get_column_letter(index)].width = width

    anomaly_sheet = workbook.create_sheet("异常清单")
    anomaly_sheet.append(["来源", "问题关键字", "异常类型", "说明"])
    for item in result["anomalies"]:
        anomaly_sheet.append([item.get("source", ""), item.get("issue_key", ""), item.get("label", ""), item.get("detail", "")])
    notes_sheet = workbook.create_sheet("处理说明")
    notes_sheet.append(["项目", "内容"])
    notes_sheet.append(["报告日期", result.get("report_date", "")])
    notes_sheet.append(["完成规则", "解决结果为空不计入完成或修复；状态与解决结果均明确完成时才计入。"])
    notes_sheet.append(["逾期规则", "交付日期早于报告日期且未完成，视为逾期。"])
    notes_sheet.append(["人员规则", "经办人与开发人员不一致时按经办人基础统计并追加开发人员；开发人员为空按经办人。"])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.title != "周报统计":
            sheet.auto_filter.ref = sheet.dimensions
    workbook.save(target)
    workbook.close()


def export_jira_xlsx(result: dict[str, Any], target: str | Path) -> None:
    """导出人员汇总、任务明细、异常清单和字段映射。"""
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    def add_sheet(title: str, headers: list[str], rows: list[list[Any]]) -> None:
        sheet = workbook.create_sheet(title)
        sheet.append([_excel_safe(value) for value in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        for row in rows:
            sheet.append([_excel_safe(value) for value in row])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 10), 42)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        if title == "人员汇总":
            sheet.column_dimensions["B"].width = 70
            for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                row[1].alignment = Alignment(vertical="top", wrap_text=True)
                line_count = max(1, str(row[1].value or "").count("\n") + 1)
                sheet.row_dimensions[row_index].height = min(18 * line_count, 120)

    add_sheet(
        "人员汇总",
        ["人员", "合并任务内容"],
        [[item["name"], item["merged_text"]] for item in result["summaries"]],
    )
    add_sheet(
        "任务明细",
        ["来源行", "问题关键字", "问题ID", "父级ID", "原始状态", "展示状态", "创建日期", "概要", "经办人", "报告人", "开发人员", "汇总人员", "汇总依据", "备注", "模块", "交付日期", "解决结果", "问题类型", "计划完成日期", "一致性", "是否纳入汇总"],
        [[record[field] if field != "是否纳入汇总" else ("是" if record["included"] else "否") for field in ["source_row", "issue_key", "issue_id", "parent_id", "status_raw", "status", "created_at", "summary", "assignee", "reporter", "developer", "summary_person", "summary_source", "summary_note", "module", "delivery_date", "resolution", "issue_type", "planned_completion", "consistency", "是否纳入汇总"]] for record in result["records"]],
    )
    add_sheet(
        "异常清单",
        ["来源行", "问题关键字", "异常类型", "说明"],
        [[item["row"], item["issue_key"], item["label"], item["detail"]] for item in result["anomalies"]],
    )
    add_sheet(
        "字段映射",
        ["标准字段", "原始表头", "列", "是否识别"],
        [[field, details["header"], details["column"], "是" if details["found"] else "否"] for field, details in result["field_mapping"].items()],
    )
    workbook.save(target)
    workbook.close()


def export_daily_jira_xlsx(result: dict[str, Any], target: str | Path) -> None:
    """导出每日 Jira 最终人员汇总和人员归属明细。"""
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "每日Jira汇总"
    summary_sheet.append(["汇总人员", "合并内容"])
    for item in result["summaries"]:
        summary_sheet.append([_excel_safe(item["name"]), _excel_safe(item["merged_text"])])

    detail_sheet = workbook.create_sheet("筛选明细")
    detail_sheet.append(["来源行", "问题关键字", "开发人员", "经办人", "报告人", "汇总人员", "汇总依据", "备注", "原始状态", "转换状态", "概要", "合并内容", "人员情况", "是否纳入"])
    for record in result["records"]:
        detail_sheet.append([_excel_safe(value) for value in [
            record["source_row"], record["issue_key"], record["developer"], record["assignee"],
            record["reporter"], record["summary_person"], record["summary_source"], record["summary_note"],
            record["status_raw"], record["status"], record["summary"], record["summary_text"],
            record["consistency"], "是" if record["included"] else "否",
        ]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in (summary_sheet, detail_sheet):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary_sheet.column_dimensions["A"].width = 18
    summary_sheet.column_dimensions["B"].width = 80
    for row_index in range(2, summary_sheet.max_row + 1):
        line_count = max(1, str(summary_sheet.cell(row_index, 2).value or "").count("\n") + 1)
        summary_sheet.row_dimensions[row_index].height = min(18 * line_count, 180)
    detail_widths = [10, 18, 18, 18, 18, 18, 14, 40, 14, 14, 42, 52, 24, 12]
    for index, width in enumerate(detail_widths, start=1):
        detail_sheet.column_dimensions[get_column_letter(index)].width = width
    workbook.save(target)
    workbook.close()
